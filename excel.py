from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment


class Starting:

    """
    Класс для создание таблицы
    Starting:
        создание excel
        __init__:
            lin: список для информации в ячейках
            num: сколько всего должно быть рядов
        center:
            row: ячейка ряда
            column: ячейка столбца
            text: текст в ячейки
            "
            задает размер текста
            центрирует текст
            установка текста в ячейк
            "
        update:
        отрисовка таблицы

        первый цикл записывает превый ряд

        второй отрисовывет таблицу по заданным данным self.info: lin и self.number: num

        третий задает ширину столбца


    """

    def __init__(self, lin: list, num: int):
        self.number = num
        self.info = lin
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "taxi"
        self.row = {'A': 10, 'B': 15, 'C': 20, 'D': 20, 'E': 25, 'F': 25}
        self.ws.row_dimensions[1].height = 30

    def center(self, row: int, column: int, text: str) -> bool:
        self.ws.cell(row=row, column=column).font = Font(size=13, bold=True)
        self.ws.cell(row=row, column=column).alignment = Alignment(horizontal='center', vertical='center')
        self.ws.cell(row=row, column=column).value = f' {text} '

    def update(self):
        self.sum = 0
        text_ = ["X", "Время", "Имя водителя",
                 "Имя пассажира", "Пункт отправления", "Пункт назначения"]
        for i in range(0, 6):
            self.center(row=1, column=i + 1, text=text_[i])

        for i in range(0, self.number):
            self.ws.row_dimensions[i + 2].height = 40
            self.ws.cell(row=i + 2, column=1).font = Font(size=13, bold=True)
            self.center(row=i + 2, column=1, text=i + 1)
            for j in range(5):
                self.ws.cell(row=i + 2, column=j + 2).font = Font(size=13, bold=True)
                self.center(row=i + 2, column=j + 2, text=self.info[int(self.sum)])
                self.sum += 1

        for i in self.row.keys():
            self.ws.column_dimensions[i].width = self.row[i]
        self.wb.save('taxi.xlsx')

